from __future__ import annotations

import argparse
import copy
import json
import re
import sys
from dataclasses import dataclass
from datetime import datetime, time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.styles import Font, PatternFill


DEFAULT_DATA_SHEET = "2026 Data"
SPRINKLR_HEADER_ROW = 3
MASTER_HEADER_ROW = 1
DELETED_POST_TOKEN = "deleted post"


SPRINKLR_TO_MASTER = {
    "Published Date": "Date",
    "PublishedTime": "Time",
    "Outbound Post": "Outbound Post",
    "Campaign Name": "Campaign Name",
    "Permalink": "Permalink",
    "YT - Theme": "YT Theme",
    "IG Post Reach": "Reach",
    "IG Post Views": "Views",
    "[YT] Positive Engagements (SUM)": "PE",
    "IG Likes": "Likes",
    "IG Saves": "Saves",
    "IG Shares": "Shares",
    "IG Reposts": "Reposts",
    "IG Reels Skip Rate": "Skip Rate %",
    "IG Reels Average Watch Time": "Avg Watch Time (Seconds)",
    "IG Share Rate": "Share rate",
    "IG Comments": "Comments",
    "IG Positive Comments": "Pos Comments",
    "IG Negative Comments": "Neg Comments",
    "IG Neutral Comments": "Neutral Comments",
    "[SLI] Engagements (SUM)": "Vis Engagement",
    "[SLI] Engagement Rate in %": "ER%",
    "IG Post Followers Gained": "Follows",
}

PERCENT_AS_DECIMAL_SOURCE_COLUMNS = {
    "IG Share Rate",
    "[SLI] Engagement Rate in %",
}

NUMERIC_SOURCE_COLUMNS = {
    "IG Post Reach",
    "IG Post Views",
    "[YT] Positive Engagements (SUM)",
    "IG Likes",
    "IG Saves",
    "IG Shares",
    "IG Reposts",
    "IG Reels Skip Rate",
    "IG Comments",
    "IG Positive Comments",
    "IG Negative Comments",
    "IG Neutral Comments",
    "[SLI] Engagements (SUM)",
    "IG Post Followers Gained",
}

WEEKLY_FORMULA_COLUMNS = {
    "Year",
    "H1/H2",
    "Quarter",
    "Month",
    "Fortnight",
    "Week",
    "Day",
    "Hourly Buckets",
    "Topic_List",
    "Topic_Options",
}

PERIODICAL_UPDATE_COLUMNS = {
    "Reach",
    "Views",
    "PE",
    "Likes",
    "Saves",
    "Shares",
    "Reposts",
    "Skip Rate %",
    "Avg Watch Time (Seconds)",
    "Share rate",
    "Comments",
    "Pos Comments",
    "Neg Comments",
    "Neutral Comments",
    "Vis Engagement",
    "ER%",
    "Follows",
}

EXCEL_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm"}


@dataclass
class RunStats:
    mode: str
    master_path: str = ""
    sprinklr_path: str = ""
    output_path: str = ""
    sprinklr_rows_seen: int = 0
    deleted_posts_removed: int = 0
    missing_permalink_rows: int = 0
    duplicate_rows_skipped: int = 0
    rows_appended: int = 0
    rows_updated: int = 0
    unmatched_periodical_rows: int = 0
    missing_source_columns: list[str] | None = None
    missing_master_columns: list[str] | None = None
    row_audit: list[dict[str, Any]] | None = None


def default_config() -> dict[str, Any]:
    return {
        "data_sheet": DEFAULT_DATA_SHEET,
        "sprinklr_header_row": SPRINKLR_HEADER_ROW,
        "master_header_row": MASTER_HEADER_ROW,
        "deleted_post_token": DELETED_POST_TOKEN,
        "master_path": "master/Google sheet sample.xlsx",
        "weekly_input_dir": "input/weekly",
        "periodical_input_dir": "input/periodical",
        "output_dir": "outputs",
        "log_dir": "logs",
        "fail_on_missing_source_columns": True,
        "sprinklr_to_master": SPRINKLR_TO_MASTER,
        "periodical_update_columns": sorted(PERIODICAL_UPDATE_COLUMNS),
        "percent_as_decimal_source_columns": sorted(PERCENT_AS_DECIMAL_SOURCE_COLUMNS),
        "numeric_source_columns": sorted(NUMERIC_SOURCE_COLUMNS),
        "reel_permalink_contains": "reel",
        "non_reel_override_columns": {
            "IG Reels Skip Rate": "NA",
            "IG Reels Average Watch Time": "NA",
        },
    }


def load_config(path: Path | None) -> dict[str, Any]:
    config = default_config()
    if path and path.exists():
        incoming = json.loads(path.read_text(encoding="utf-8"))
        config.update(incoming)
    return config


def resolve_path(path_value: str | Path, base_dir: Path) -> Path:
    path = Path(path_value)
    return path if path.is_absolute() else base_dir / path


def latest_excel_file(input_dir: Path) -> Path:
    candidates = [
        path
        for path in input_dir.iterdir()
        if path.is_file() and path.suffix.lower() in EXCEL_EXTENSIONS and not path.name.startswith("~$")
    ]
    if not candidates:
        raise FileNotFoundError(f"No Excel files found in {input_dir}")
    return max(candidates, key=lambda path: path.stat().st_mtime)


def safe_stem(path: Path) -> str:
    return re.sub(r"[^A-Za-z0-9._-]+", "_", path.stem).strip("_") or "sprinklr_dump"


def audit_entry(
    source_row: int,
    action: str,
    permalink: Any,
    reason: str,
    target_row: int | None = None,
) -> dict[str, Any]:
    return {
        "source_row": source_row,
        "action": action,
        "permalink": normalize_permalink(permalink),
        "reason": reason,
        "target_row": target_row,
    }


def normalize_header(value: Any) -> str:
    return "" if value is None else str(value).strip()


def normalize_permalink(value: Any) -> str:
    return "" if value is None else str(value).strip()


def parse_datetime(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        text = value.strip()
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M:%S.%f", "%Y-%m-%d"):
            try:
                return datetime.strptime(text, fmt)
            except ValueError:
                pass
    return None


def split_date(value: Any) -> datetime | Any:
    dt = parse_datetime(value)
    return dt.date() if dt else value


def split_time(value: Any) -> time | Any:
    dt = parse_datetime(value)
    return dt.time().replace(microsecond=0) if dt else value


def coerce_number(value: Any) -> Any:
    if value is None or isinstance(value, (int, float)):
        return value
    if isinstance(value, str):
        text = value.strip().replace(",", "")
        if text == "":
            return value
        try:
            number = float(text)
            return int(number) if number.is_integer() else number
        except ValueError:
            return value
    return value


def percent_to_decimal(value: Any) -> Any:
    if value is None or isinstance(value, (int, float)):
        return value
    if isinstance(value, str):
        text = value.strip().replace(",", "")
        if text.endswith("%"):
            try:
                return float(text[:-1]) / 100
            except ValueError:
                return value
    return coerce_number(value)


def is_reel_permalink(value: Any, reel_token: str = "reel") -> bool:
    permalink = normalize_permalink(value).lower()
    return reel_token.lower() in permalink


def is_deleted_post(row: dict[str, Any], deleted_post_token: str = DELETED_POST_TOKEN) -> bool:
    campaign_values = [
        row.get("Campaign"),
        row.get("Campaign Name"),
    ]
    return any(deleted_post_token.lower() in str(value).lower() for value in campaign_values if value is not None)


def header_map(ws, header_row: int) -> dict[str, int]:
    result: dict[str, int] = {}
    for cell in ws[header_row]:
        header = normalize_header(cell.value)
        if header:
            result[header] = cell.column
    return result


def read_sprinklr_rows(path: Path, header_row: int = SPRINKLR_HEADER_ROW) -> list[dict[str, Any]]:
    wb = load_workbook(path, data_only=False, read_only=True)
    ws = wb.active
    headers = [normalize_header(cell.value) for cell in ws[header_row]]
    rows: list[dict[str, Any]] = []
    for source_row, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        if not any(value is not None and str(value).strip() for value in row):
            continue
        parsed_row = {headers[i]: value for i, value in enumerate(row) if i < len(headers)}
        parsed_row["_source_row"] = source_row
        rows.append(parsed_row)
    return rows


def existing_permalink_index(ws, master_headers: dict[str, int], master_header_row: int = MASTER_HEADER_ROW) -> dict[str, int]:
    permalink_col = master_headers["Permalink"]
    index: dict[str, int] = {}
    for row_idx in range(master_header_row + 1, ws.max_row + 1):
        permalink = normalize_permalink(ws.cell(row=row_idx, column=permalink_col).value)
        if permalink:
            index[permalink] = row_idx
    return index


def first_blank_permalink_row(
    ws, master_headers: dict[str, int], master_header_row: int = MASTER_HEADER_ROW
) -> int:
    permalink_col = master_headers["Permalink"]
    for row_idx in range(master_header_row + 1, ws.max_row + 2):
        value = normalize_permalink(ws.cell(row=row_idx, column=permalink_col).value)
        if not value:
            return row_idx
    return ws.max_row + 1


def translate_formula(formula: str, origin: str, target: str) -> str:
    try:
        return Translator(formula, origin=origin).translate_formula(target)
    except Exception:
        return formula


def copy_formula_row(ws, from_row: int, to_row: int) -> None:
    if from_row == to_row:
        return
    for col_idx in range(1, ws.max_column + 1):
        source = ws.cell(row=from_row, column=col_idx)
        target = ws.cell(row=to_row, column=col_idx)
        if isinstance(source.value, str) and source.value.startswith("="):
            target.value = translate_formula(source.value, source.coordinate, target.coordinate)
        if source.has_style:
            target._style = copy.copy(source._style)
        if source.number_format:
            target.number_format = source.number_format


def target_value(source_header: str, raw_value: Any, sprinklr_row: dict[str, Any], config: dict[str, Any]) -> Any:
    non_reel_overrides = config.get("non_reel_override_columns", {})
    if source_header in non_reel_overrides and not is_reel_permalink(
        sprinklr_row.get("Permalink"), config.get("reel_permalink_contains", "reel")
    ):
        return non_reel_overrides[source_header]
    if source_header == "Published Date":
        return split_date(raw_value)
    if source_header == "PublishedTime":
        return split_time(raw_value)
    if source_header in set(config.get("percent_as_decimal_source_columns", [])):
        return percent_to_decimal(raw_value)
    if source_header in set(config.get("numeric_source_columns", [])):
        return coerce_number(raw_value)
    return raw_value


def transform_sprinklr_row(sprinklr_row: dict[str, Any], config: dict[str, Any]) -> dict[str, Any]:
    transformed: dict[str, Any] = {}
    for source_header, target_header in config["sprinklr_to_master"].items():
        transformed[target_header] = target_value(
            source_header,
            sprinklr_row.get(source_header),
            sprinklr_row,
            config,
        )
    return transformed


def write_mapped_row(
    ws, row_idx: int, master_headers: dict[str, int], sprinklr_row: dict[str, Any], config: dict[str, Any]
) -> None:
    published_date = split_date(sprinklr_row.get("Published Date"))
    if "Year" in master_headers and hasattr(published_date, "year"):
        ws.cell(row=row_idx, column=master_headers["Year"]).value = published_date.year

    for source_header, target_header in config["sprinklr_to_master"].items():
        if target_header not in master_headers:
            continue
        ws.cell(row=row_idx, column=master_headers[target_header]).value = target_value(
            source_header,
            sprinklr_row.get(source_header),
            sprinklr_row,
            config,
        )


def validate_columns(
    master_headers: dict[str, int], sprinklr_rows: list[dict[str, Any]], config: dict[str, Any]
) -> tuple[list[str], list[str]]:
    source_headers = set(sprinklr_rows[0].keys()) if sprinklr_rows else set()
    mapped = config["sprinklr_to_master"]
    missing_source = sorted(source for source in mapped if source not in source_headers)
    needed_master = set(mapped.values()) | {"Permalink"} | set(config.get("periodical_update_columns", []))
    missing_master = sorted(header for header in needed_master if header not in master_headers)
    return missing_source, missing_master


def reset_sheet(wb, title: str):
    if title in wb.sheetnames:
        del wb[title]
    return wb.create_sheet(title)


def style_header(row) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in row:
        cell.fill = fill
        cell.font = font


def add_validation_report(wb, stats: RunStats) -> None:
    summary = reset_sheet(wb, "Run Summary")
    summary_rows = [
        ("Metric", "Value"),
        ("Mode", stats.mode),
        ("Master path", stats.master_path),
        ("Sprinklr path", stats.sprinklr_path),
        ("Output path", stats.output_path),
        ("Sprinklr rows seen", stats.sprinklr_rows_seen),
        ("Deleted posts removed", stats.deleted_posts_removed),
        ("Missing permalink rows", stats.missing_permalink_rows),
        ("Duplicate rows skipped", stats.duplicate_rows_skipped),
        ("Rows appended", stats.rows_appended),
        ("Rows updated", stats.rows_updated),
        ("Unmatched periodical rows", stats.unmatched_periodical_rows),
        ("Missing source columns", ", ".join(stats.missing_source_columns or [])),
        ("Missing master columns", ", ".join(stats.missing_master_columns or [])),
    ]
    for row in summary_rows:
        summary.append(row)
    style_header(summary[1])
    summary.column_dimensions["A"].width = 28
    summary.column_dimensions["B"].width = 110

    audit = reset_sheet(wb, "Row Audit")
    audit_headers = ["Source Row", "Action", "Permalink", "Reason", "Target Row"]
    audit.append(audit_headers)
    for entry in stats.row_audit or []:
        audit.append(
            [
                entry.get("source_row"),
                entry.get("action"),
                entry.get("permalink"),
                entry.get("reason"),
                entry.get("target_row"),
            ]
        )
    style_header(audit[1])
    audit.column_dimensions["A"].width = 12
    audit.column_dimensions["B"].width = 18
    audit.column_dimensions["C"].width = 65
    audit.column_dimensions["D"].width = 42
    audit.column_dimensions["E"].width = 12
    audit.freeze_panes = "A2"


def run_weekly(master_path: Path, sprinklr_path: Path, output_path: Path, data_sheet: str, config: dict[str, Any]) -> RunStats:
    stats = RunStats(mode="weekly_append")
    stats.row_audit = []
    stats.master_path = str(master_path)
    stats.sprinklr_path = str(sprinklr_path)
    stats.output_path = str(output_path)
    wb = load_workbook(master_path)
    ws = wb[data_sheet]
    master_header_row = int(config.get("master_header_row", MASTER_HEADER_ROW))
    sprinklr_header_row = int(config.get("sprinklr_header_row", SPRINKLR_HEADER_ROW))
    master_headers = header_map(ws, master_header_row)
    existing = existing_permalink_index(ws, master_headers, master_header_row)
    rows = read_sprinklr_rows(sprinklr_path, sprinklr_header_row)
    stats.sprinklr_rows_seen = len(rows)
    missing_source, missing_master = validate_columns(master_headers, rows, config)
    stats.missing_source_columns = missing_source
    stats.missing_master_columns = missing_master
    if missing_source and config.get("fail_on_missing_source_columns", True):
        add_validation_report(wb, stats)
        wb.save(output_path)
        raise ValueError(f"Missing Sprinklr source columns: {missing_source}. Validation report saved to {output_path}")
    if missing_master:
        raise ValueError(f"Missing master columns: {missing_master}")
    append_row = first_blank_permalink_row(ws, master_headers, master_header_row)
    formula_template_row = master_header_row + 1

    for sprinklr_row in rows:
        source_row = int(sprinklr_row.get("_source_row", 0))
        if is_deleted_post(sprinklr_row, config.get("deleted_post_token", DELETED_POST_TOKEN)):
            stats.deleted_posts_removed += 1
            stats.row_audit.append(
                audit_entry(source_row, "removed", sprinklr_row.get("Permalink"), "Deleted post")
            )
            continue
        permalink = normalize_permalink(sprinklr_row.get("Permalink"))
        if not permalink:
            stats.missing_permalink_rows += 1
            stats.row_audit.append(audit_entry(source_row, "skipped", permalink, "Missing permalink"))
            continue
        if permalink in existing:
            stats.duplicate_rows_skipped += 1
            stats.row_audit.append(
                audit_entry(source_row, "skipped", permalink, "Permalink already exists", existing[permalink])
            )
            continue

        copy_formula_row(ws, formula_template_row, append_row)
        write_mapped_row(ws, append_row, master_headers, sprinklr_row, config)
        stats.row_audit.append(audit_entry(source_row, "appended", permalink, "New permalink", append_row))
        existing[permalink] = append_row
        stats.rows_appended += 1
        append_row += 1

    add_validation_report(wb, stats)
    wb.save(output_path)
    return stats


def run_periodical(
    master_path: Path, sprinklr_path: Path, output_path: Path, data_sheet: str, config: dict[str, Any]
) -> RunStats:
    stats = RunStats(mode="periodical_update")
    stats.row_audit = []
    stats.master_path = str(master_path)
    stats.sprinklr_path = str(sprinklr_path)
    stats.output_path = str(output_path)
    wb = load_workbook(master_path)
    ws = wb[data_sheet]
    master_header_row = int(config.get("master_header_row", MASTER_HEADER_ROW))
    sprinklr_header_row = int(config.get("sprinklr_header_row", SPRINKLR_HEADER_ROW))
    master_headers = header_map(ws, master_header_row)
    existing = existing_permalink_index(ws, master_headers, master_header_row)
    rows = read_sprinklr_rows(sprinklr_path, sprinklr_header_row)
    stats.sprinklr_rows_seen = len(rows)
    missing_source, missing_master = validate_columns(master_headers, rows, config)
    stats.missing_source_columns = missing_source
    stats.missing_master_columns = missing_master
    if missing_source and config.get("fail_on_missing_source_columns", True):
        add_validation_report(wb, stats)
        wb.save(output_path)
        raise ValueError(f"Missing Sprinklr source columns: {missing_source}. Validation report saved to {output_path}")
    if missing_master:
        raise ValueError(f"Missing master columns: {missing_master}")

    reverse_map = {target: source for source, target in config["sprinklr_to_master"].items()}
    for sprinklr_row in rows:
        source_row = int(sprinklr_row.get("_source_row", 0))
        if is_deleted_post(sprinklr_row, config.get("deleted_post_token", DELETED_POST_TOKEN)):
            stats.deleted_posts_removed += 1
            stats.row_audit.append(
                audit_entry(source_row, "removed", sprinklr_row.get("Permalink"), "Deleted post")
            )
            continue
        permalink = normalize_permalink(sprinklr_row.get("Permalink"))
        if not permalink:
            stats.missing_permalink_rows += 1
            stats.row_audit.append(audit_entry(source_row, "skipped", permalink, "Missing permalink"))
            continue
        row_idx = existing.get(permalink)
        if row_idx is None:
            stats.unmatched_periodical_rows += 1
            stats.row_audit.append(audit_entry(source_row, "unmatched", permalink, "Permalink not found"))
            continue

        for target_header in set(config.get("periodical_update_columns", [])):
            if target_header not in master_headers:
                continue
            source_header = reverse_map.get(target_header)
            if source_header:
                ws.cell(row=row_idx, column=master_headers[target_header]).value = target_value(
                    source_header,
                    sprinklr_row.get(source_header),
                    sprinklr_row,
                    config,
                )
        stats.rows_updated += 1
        stats.row_audit.append(audit_entry(source_row, "updated", permalink, "Matched permalink", row_idx))

    add_validation_report(wb, stats)
    wb.save(output_path)
    return stats


def main() -> None:
    base_dir = Path(__file__).resolve().parent
    parser = argparse.ArgumentParser(description="Process Sprinklr dumps against the Instagram master sheet.")
    parser.add_argument("--mode", choices=["weekly", "periodical"], required=True)
    parser.add_argument("--config", default=base_dir / "config" / "settings.json", type=Path)
    parser.add_argument("--master", type=Path)
    parser.add_argument("--sprinklr", type=Path)
    parser.add_argument("--input-dir", type=Path)
    parser.add_argument("--output", type=Path)
    parser.add_argument("--output-dir", type=Path)
    parser.add_argument("--data-sheet")
    parser.add_argument("--stats-json", type=Path)
    args = parser.parse_args()

    config = load_config(args.config)
    data_sheet = args.data_sheet or config.get("data_sheet", DEFAULT_DATA_SHEET)
    master_path = args.master or resolve_path(config["master_path"], base_dir)
    if args.sprinklr:
        sprinklr_path = args.sprinklr
    else:
        input_dir_key = "weekly_input_dir" if args.mode == "weekly" else "periodical_input_dir"
        input_dir = args.input_dir or resolve_path(config[input_dir_key], base_dir)
        sprinklr_path = latest_excel_file(input_dir)

    output_dir = args.output_dir or resolve_path(config.get("output_dir", "outputs"), base_dir)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    if args.output:
        output_path = args.output
    else:
        output_name = f"{args.mode}_{safe_stem(sprinklr_path)}_{timestamp}.xlsx"
        output_path = output_dir / output_name

    log_dir = resolve_path(config.get("log_dir", "logs"), base_dir)
    stats_json = args.stats_json or log_dir / f"{args.mode}_{safe_stem(sprinklr_path)}_{timestamp}.json"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        if args.mode == "weekly":
            stats = run_weekly(master_path, sprinklr_path, output_path, data_sheet, config)
        else:
            stats = run_periodical(master_path, sprinklr_path, output_path, data_sheet, config)
    except Exception as exc:
        payload = {
            "mode": args.mode,
            "master_path": str(master_path),
            "sprinklr_path": str(sprinklr_path),
            "output_path": str(output_path),
            "status": "failed",
            "error": str(exc),
        }
        print(json.dumps(payload, indent=2))
        stats_json.parent.mkdir(parents=True, exist_ok=True)
        stats_json.write_text(json.dumps(payload, indent=2), encoding="utf-8")
        sys.exit(1)

    payload = stats.__dict__
    payload["status"] = "success"
    print(json.dumps(payload, indent=2))
    stats_json.parent.mkdir(parents=True, exist_ok=True)
    stats_json.write_text(json.dumps(payload, indent=2), encoding="utf-8")


if __name__ == "__main__":
    main()
